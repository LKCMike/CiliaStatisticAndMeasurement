"""
:Date        : 2025-07-22 16:48:15
:LastEditTime: 2025-08-28 10:03:44
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtCore import QObject, QDir, QFileInfo, QUrl
from PyQt5.QtGui import QDesktopServices
from ultralytics import YOLO
import cv2
from measurement import arc_length, get_contour_size
from misc import parse_json


font = Font(name='微软雅黑', size=11)
alignment = Alignment(horizontal='right', vertical='center')

class MainController(QObject):
    """
    UI控件功能定义
    """
    def __init__(self, view):
        super().__init__()
        self.view = view
        self.setup_connections()

    def setup_connections(self):
        """连接所有信号与槽"""
        # 文件操作按钮
        self.view.add_btn.clicked.connect(self.add_selected_files)
        self.view.remove_btn.clicked.connect(self.remove_selected_files)
        self.view.start_btn.clicked.connect(self.start_processing)

        self.view.output_dir_btn.clicked.connect(self.select_output_dir)
        # 连接打开文件夹按钮
        self.view.open_dir_btn.clicked.connect(self.open_output_directory)

    def open_output_directory(self):
        """打开输出文件夹"""
        dir_path = self.view.output_dir_input.text().strip()

        if not dir_path:
            self.view.append_log("[Error]: Please choose output directory first.")
            return

        if not QDir(dir_path).exists():
            self.view.append_log(f"[Error]: No such directory - {dir_path}")
            return

        # 使用系统默认文件管理器打开文件夹
        success = QDesktopServices.openUrl(QUrl.fromLocalFile(dir_path))

        if not success:
            self.view.append_log(f"[Error] Unable to open directory in File Explorer: {dir_path}")

    def add_selected_files(self):
        """添加选中的文件到右侧列表"""
        selected = self.view.tree_view.selectedIndexes()
        if not selected:
            self.view.append_log("[Warning]: No file was chosen to add to queue.")
            return

        for index in selected:
            if self.view.file_model.isDir(index):
                continue
            file_path = self.view.file_model.filePath(index)
            if file_path not in [self.view.selected_files_list.item(i).text()
                               for i in range(self.view.selected_files_list.count())]:
                self.view.selected_files_list.addItem(file_path)
                self.view.append_log(f"[Log] File Appended: {file_path}")

    def remove_selected_files(self):
        """从右侧列表移除选中的文件"""
        selected = self.view.selected_files_list.selectedItems()
        if not selected:
            self.view.append_log("[Warning]: No File was chosen to remove from queue.")
            return

        for item in selected:
            self.view.append_log(f"[Log] File deleted: {item.text()}")
            self.view.selected_files_list.takeItem(
                self.view.selected_files_list.row(item))

    def select_output_dir(self):
        """选择输出目录"""
        dir_path = QFileDialog.getExistingDirectory(
            self.view, "Output", QDir.homePath())
        if dir_path:
            self.view.output_dir_input.setText(dir_path)
            self.view.append_log(f"[Log] Set Output Directory: {dir_path}")

    def start_processing(self):
        """开始处理文件"""
        model_conf = parse_json("config.json")
        if self.view.selected_files_list.count() == 0:
            self.view.append_log("[Error]: No File to Process")
            return

        if not self.view.output_dir_input.text():
            self.view.append_log("[Error]: Please Choose Your Output Directory.")
            return

        # 获取所有选中的文件
        files = [self.view.selected_files_list.item(i).text()
                for i in range(self.view.selected_files_list.count())]
        output_dir = self.view.output_dir_input.text().strip()

        self.view.append_log("[Log] Start Processing...")
        self.view.append_log(f"     Files in Queue: {len(files)}")
        self.view.append_log(f"     Output to: {output_dir}")

        model = YOLO("model/cilia-detection.pt")
        tag_list = []
        for file in files:
            try:
                contour_size = get_contour_size(file)
            except FileNotFoundError:
                self.view.append_log(f"[Error] No metadata XML File was found for {file}. Unable to measure cilia length.")
                contour_size = None
            except ValueError as e:
                self.view.append_log(f"[Error] Invalid value found in XML. ContourSize should be int or float. {repr(e)}")
                contour_size = None
            img = cv2.imread(file)
            file : str
            file_info = QFileInfo(file)
            base_name = file_info.baseName()
            tag_output_path = os.path.join(output_dir, f"{base_name}.xlsx")
            img_output_path = os.path.join(output_dir, f"{base_name}.png")
            tag_list.clear()
            self.view.append_log(f"[Log] Processing: {file}")
            results = model.predict(
                source=file,
                conf=model_conf["Minimal Confidence"],     # 提高置信度阈值
                iou=model_conf["Intersection over Union Threshold"],      # 降低IoU阈值
                save=False,
                project=output_dir,  # 绝对路径
                name="",
                exist_ok=True,
                show_labels=False,  # 不显示类别标签
                show_conf=False,    # 不显示置信度
                line_width=2        # 可选：调整框线粗细
            )
            for result in results:
                output_img = result.orig_img.copy()
                boxes = result.boxes    # 包含坐标、置信度、类别
                dropcase_count = 0
                for index_id, box in enumerate(boxes, start=1):
                    box_list = box.xyxy[0].tolist()
                    for i, coordinate in enumerate(box_list):
                        box_list[i] = int(coordinate)
                    x1, y1, x2, y2 = box_list
                    classid = int(box.cls)
                    height, width, _ = img.shape
                    cilia_length = arc_length(
                        img,
                        box_list,
                        contour_size,
                        model_conf.get("Brightness Decay Limit", 0)
                    ) if classid == 1 else 0
                    if any(i in box_list for i in (0, height, width)):
                        if not int(cilia_length):
                            dropcase_count += 1
                            continue
                    rectangle_color = (0, 255, 0) if classid else (0, 0, 255)
                    cv2.rectangle(output_img, (x1, y1), (x2, y2), rectangle_color, 2)
                    label = str(index_id - dropcase_count)
                    label_offset = x1 if classid else x2 - 20
                    label_color = (255, 255, 255) if classid else (0, 255, 255)
                    if y1 - 28 > 0:
                        cv2.putText(output_img, label, (label_offset, y1 - 8), cv2.FONT_HERSHEY_SIMPLEX, 0.8, label_color, 2)
                    else:
                        cv2.putText(output_img, label, (label_offset, y1 + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.8, label_color, 2)
                    classname = "Y" if classid else "N"
                    tag_list.append([index_id - dropcase_count, classname, str(box_list), cilia_length])
            cv2.imwrite(img_output_path, output_img)
            template_excel = load_workbook("template/template.xlsx")
            template_sheet = template_excel.active
            for row_idx, row_data in enumerate(tag_list, start=4):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = template_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    cell.font = font
                    cell.alignment = alignment
            if os.path.exists(tag_output_path):
                os.remove(tag_output_path)
            template_excel.save(tag_output_path)
            template_excel.close()

        self.view.append_log(f"[Log] Finished! Check Output Directory: {output_dir}")
